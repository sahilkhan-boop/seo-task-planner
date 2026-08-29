"""Renders the campaign's month-by-month calendar to an .xlsx workbook --
one sheet per month, laid out as a calendar grid (not a flat row list), each
day cell showing its tasks with project + assignee. Mirrors the PDF/HTML views.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.scheduling.calendar_grid import WEEKDAY_LABELS, MonthGrid

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
OUT_OF_MONTH_FILL = PatternFill("solid", fgColor="F3F4F6")
SEVERITY_FILL = {
    "high": PatternFill("solid", fgColor="FCE7E7"),
    "medium": PatternFill("solid", fgColor="FDF0DE"),
    "low": PatternFill("solid", fgColor="E1F3EB"),
}
THIN_BORDER = Border(*(Side(style="thin", color="E5E7EB"),) * 4)
MAX_TASKS_PER_CELL = 6


def _day_cell_text(cell, project_name: str) -> str:
    lines = [str(cell.date.day)]
    shown = cell.tasks[:MAX_TASKS_PER_CELL]
    for t in shown:
        title = t.title if len(t.title) <= 60 else t.title[:57] + "..."
        lines.append(f"[{t.severity.upper()}] {title}")
        lines.append(f"  {project_name} · {t.assignee or 'Unassigned'}")
    if len(cell.tasks) > MAX_TASKS_PER_CELL:
        lines.append(f"+{len(cell.tasks) - MAX_TASKS_PER_CELL} more")
    return "\n".join(lines)


def _dominant_severity(cell) -> str | None:
    severities = {t.severity for t in cell.tasks}
    for sev in ("high", "medium", "low"):
        if sev in severities:
            return sev
    return None


def _safe_sheet_name(label: str, used: set[str]) -> str:
    base = label[:31]
    name = base
    n = 2
    while name in used:
        suffix = f" ({n})"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def build_calendar_xlsx(project_name: str, months: list[MonthGrid]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    for month in months:
        ws = wb.create_sheet(_safe_sheet_name(month.label, used_names))
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"{month.label} — {project_name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        for col, label in enumerate(WEEKDAY_LABELS, start=1):
            c = ws.cell(row=2, column=col, value=label)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER

        row = 3
        for week in month.weeks:
            for col, day_cell in enumerate(week, start=1):
                c = ws.cell(row=row, column=col, value=_day_cell_text(day_cell, project_name))
                c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                c.border = THIN_BORDER
                if not day_cell.in_month:
                    c.fill = OUT_OF_MONTH_FILL
                else:
                    dominant = _dominant_severity(day_cell)
                    if dominant:
                        c.fill = SEVERITY_FILL[dominant]
            ws.row_dimensions[row].height = 90
            row += 1

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 24

    if not months:
        ws = wb.create_sheet("No campaign")
        ws["A1"] = "No campaign configured yet -- set one up in Settings."

    from io import BytesIO

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
